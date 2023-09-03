# Look for unusual network activity
# Judgment by checking whether the number of DNS queries of data packets per week has changed within four weeks

import os
from scapy.all import *
from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter

def extract_dns_queries_from_pcap(file_path):
    packets = rdpcap(file_path)
    dns_queries = [packet[DNS].qd.qname.decode('utf-8').rstrip('.') for packet in packets if packet.haslayer(DNS) and packet[DNS].qd]
    return dns_queries

def analyze_pcap_files(folder_path):
    weekly_dns_counts = {
        "Week1": Counter(),
        "Week2": Counter(),
        "Week3": Counter(),
        "Week4": Counter()
    }

    pcap_files = [f for f in os.listdir(folder_path) if f.endswith(".pcap")]
    pcap_files.sort()

    for idx, filename in enumerate(pcap_files[:40]):
        dns_queries = extract_dns_queries_from_pcap(os.path.join(folder_path, filename))
        if idx < 10:
            weekly_dns_counts["Week1"].update(dns_queries)
        elif idx < 20:
            weekly_dns_counts["Week2"].update(dns_queries)
        elif idx < 30:
            weekly_dns_counts["Week3"].update(dns_queries)
        else:
            weekly_dns_counts["Week4"].update(dns_queries)

    return weekly_dns_counts

def save_to_excel(data, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DNS Query Counts"

    # Write headers
    ws['A1'] = "Domain"
    col_idx = 2
    for week in data.keys():
        ws[get_column_letter(col_idx) + '1'] = week
        col_idx += 1

    row_idx = 2
    all_domains = set(domain for week_counts in data.values() for domain in week_counts.keys())
    for domain in all_domains:
        ws[f'A{row_idx}'] = domain
        col_idx = 2
        for week in data.keys():
            ws[get_column_letter(col_idx) + str(row_idx)] = data[week].get(domain, 0)
            col_idx += 1
        row_idx += 1

    wb.save(output_file)

folder_path = "/Users/liujia/Desktop/data/wyze_cam_pan_v2"
output_excel = "/Users/liujia/Desktop/data/wyze_cam_pan_v2/dns_query_counts.xlsx"

weekly_data = analyze_pcap_files(folder_path)
save_to_excel(weekly_data, output_excel)

print(f"Results saved to {output_excel}")
